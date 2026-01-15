"""
API endpoint for data auditing
Provides comprehensive data validation and integrity checking
"""
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional, List
import os
import psycopg2
from psycopg2.extras import RealDictCursor
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.services.product_catalog_service import ProductCatalogService

router = APIRouter()

# Database connection
DATABASE_URL = os.getenv("DATABASE_URL")

# Initialize Product Catalog Service (replaces CSV dependency)
_product_catalog_service = None

def get_product_catalog_service() -> ProductCatalogService:
    """Get or initialize the ProductCatalogService singleton"""
    global _product_catalog_service
    if _product_catalog_service is None:
        _product_catalog_service = ProductCatalogService()
    return _product_catalog_service

# Load product mapping from database (replaces CSV dependency)
def load_product_mapping():
    """
    Load product catalog from database via ProductCatalogService.

    This replaces the CSV-based approach with database queries.
    Railway doesn't have access to the CSV file, so we must use the database.

    Returns: (product_map, catalog_skus, catalog_master_skus, conversion_map)
    """
    product_map = {}
    catalog_skus = set()
    catalog_master_skus = set()
    conversion_map = {}

    try:
        # Get product catalog from database
        service = get_product_catalog_service()
        all_products = service.get_all_products()

        # Build product map and catalog sets
        for sku, product_data in all_products.items():
            # Add to catalog SKUs
            catalog_skus.add(sku)

            # Check if product has a master box SKU, and add that master SKU to the set
            # This allows us to map orders that contain master box SKUs (e.g., BAMC_C02810)
            if product_data.get('sku_master'):
                catalog_master_skus.add(product_data['sku_master'])

            # Get conversion factor
            # Normal SKUs use units_per_display, master SKUs use items_per_master_box
            if product_data.get('is_master_sku'):
                conversion_factor = product_data.get('items_per_master_box', 1) or 1
            else:
                conversion_factor = product_data.get('units_per_display', 1) or 1

            # Get SKU Primario
            sku_primario = product_data.get('primario', sku)

            # Build product map entry
            product_map[sku] = {
                'category': product_data.get('category', ''),
                'family': product_data.get('product_name', ''),
                'format': 'CAJA MASTER' if product_data.get('is_master_sku') else product_data.get('format', ''),
                'product_name': product_data.get('product_name', ''),
                'in_catalog': True,
                'tipo': 'Caja Master' if product_data.get('is_master_sku') else 'Normal',
                'conversion_factor': conversion_factor,
                'sku_primario': sku_primario
            }

            # Build conversion map entry
            base_code = product_data.get('base_code', sku.split('_')[0] if '_' in sku else sku)
            conversion_map[sku] = {
                'factor': conversion_factor,
                'primario': sku_primario,
                'base_code': base_code
            }

            # ALSO add master box SKU to product_map (if it exists)
            # This allows map_sku_with_quantity() to find master SKUs directly
            # Example: order has BAMC_C02810 → needs product_map['BAMC_C02810']
            if product_data.get('sku_master'):
                master_sku = product_data['sku_master']
                # Use items_per_master_box as conversion factor for master SKUs
                master_conversion_factor = product_data.get('items_per_master_box', 1) or 1

                product_map[master_sku] = {
                    'category': product_data.get('category', ''),
                    'family': product_data.get('product_name', ''),
                    'format': 'CAJA MASTER',
                    'product_name': product_data.get('product_name', ''),
                    'in_catalog': True,
                    'tipo': 'Caja Master',
                    'conversion_factor': master_conversion_factor,
                    'sku_primario': sku_primario
                }

                # Also add to conversion map
                conversion_map[master_sku] = {
                    'factor': master_conversion_factor,
                    'primario': sku_primario,
                    'base_code': base_code
                }

        print(f"✅ Loaded {len(product_map)} products from database (includes master box SKUs)")

    except Exception as e:
        print(f"❌ Error loading product catalog from database: {e}")
        # Return empty catalogs if database fails
        pass

    return product_map, catalog_skus, catalog_master_skus, conversion_map


def get_sku_primario(sku: str, conversion_map: dict = None) -> str:
    """
    Get the SKU Primario (base/primary SKU) for any SKU variant.
    Returns the X1 variant (_U04010 for Spanish, _U04020 for English) if it exists.
    If it doesn't exist in the catalog, returns the SKU itself.

    CRITICAL: We NEVER invent SKU codes that don't exist in the catalog.

    Phase 3: Now uses ProductCatalogService (database) instead of CSV.
    The conversion_map parameter is kept for backward compatibility but is ignored.

    Example:
        BAKC_U20010 → BAKC_U04010 (primary exists in catalog)
        GRAL_U26010 → GRAL_U26010 (GRAL_U04010 doesn't exist, so use itself)
    """
    if not sku:
        return None

    # Phase 3: Use ProductCatalogService (database query)
    service = get_product_catalog_service()
    return service.get_sku_primario(sku)


def calculate_units(sku: str, quantity: int, conversion_map: dict = None, source: str = None) -> int:
    """
    Calculate total units based on SKU and quantity.

    Formula: Units = Quantity × SKU Mapping Multiplier × Target SKU Conversion Factor

    Phase 3: Now uses ProductCatalogService (database) instead of CSV.
    The conversion_map parameter is kept for backward compatibility but is ignored.

    Args:
        sku: Product SKU (original, as it appears in order)
        quantity: Quantity ordered
        conversion_map: Deprecated, ignored
        source: Data source (e.g., 'relbase') for source-specific mappings

    Examples:
        - BAKC_U04010 (X1): 10 × 1 = 10 units
        - BAKC_U20010 (X5): 10 × 5 = 50 units
        - BAKC_C02810 (CM X5): 17 × 140 = 2,380 units
        - KEEPERPACK (mapped ×5): 7 × 5 = 35 units
    """
    if not sku or quantity is None:
        return 0

    # Phase 3: Use ProductCatalogService (database query)
    service = get_product_catalog_service()
    return service.calculate_units(sku, quantity, source)


def get_pack_component_mappings(sku: str, cursor) -> list:
    """
    Get all component mappings for a PACK/variety product.

    Returns list of component dictionaries if SKU has multiple mappings (is a variety pack),
    otherwise returns empty list for normal products.

    Each component dict contains:
    - target_sku: Component SKU
    - quantity_multiplier: How many of this component in the pack
    - product_name: Name from product catalog
    - sku_primario: Primary SKU for grouping
    - category: Product category
    - sku_value: Value for proportional revenue calculation

    Example for PACKNAVIDAD2:
    [
        {'target_sku': 'KSMC_U15010', 'quantity_multiplier': 2, 'sku_value': 1886, ...},
        {'target_sku': 'BABE_U20010', 'quantity_multiplier': 1, 'sku_value': 1461, ...},
        ...
    ]
    """
    if not sku:
        return []

    cursor.execute('''
        SELECT
            sm.target_sku,
            sm.quantity_multiplier,
            pc.product_name,
            pc.sku_primario,
            pc.category,
            COALESCE(pc.sku_value, 1000) as sku_value
        FROM sku_mappings sm
        LEFT JOIN product_catalog pc ON pc.sku = sm.target_sku AND pc.is_active = TRUE
        WHERE sm.source_pattern = %s
          AND sm.pattern_type = 'exact'
          AND sm.is_active = TRUE
        ORDER BY sm.id
    ''', (sku.upper(),))

    mappings = cursor.fetchall()

    # Only return mappings if there are multiple (variety pack)
    # Single mappings are handled normally (not expanded)
    if len(mappings) <= 1:
        return []

    # Convert to list of dicts (cursor returns RealDictCursor rows)
    return [
        {
            'target_sku': row['target_sku'],
            'quantity_multiplier': row['quantity_multiplier'] or 1,
            'product_name': row['product_name'],
            'sku_primario': row['sku_primario'],
            'category': row['category'],
            'sku_value': row['sku_value'] or 1000
        }
        for row in mappings
    ]


def expand_pack_to_components(row_dict: dict, pack_mappings: list, cursor) -> list:
    """
    Expand a PACK order item into individual component rows.

    Takes a single order row containing a variety pack and returns multiple rows,
    one for each component product, with proportional revenue distribution.

    Args:
        row_dict: Original order item row (e.g., PACKNAVIDAD2, qty=1, $40,112)
        pack_mappings: List of component mappings from get_pack_component_mappings()
        cursor: Database cursor for additional lookups

    Returns:
        List of expanded row dicts, one per component
    """
    if not pack_mappings:
        return [row_dict]

    expanded_rows = []
    original_qty = row_dict.get('quantity', 0) or 0
    original_subtotal = row_dict.get('item_subtotal', 0) or 0
    original_sku = row_dict.get('sku', '')
    service = get_product_catalog_service()

    # Calculate total weighted value for proportional revenue split
    total_weighted_value = sum(
        m['sku_value'] * m['quantity_multiplier']
        for m in pack_mappings
    )

    for mapping in pack_mappings:
        component_sku = mapping['target_sku']
        qty_multiplier = mapping['quantity_multiplier']
        component_value = mapping['sku_value']
        weighted_value = component_value * qty_multiplier

        # Calculate proportional revenue
        if total_weighted_value > 0:
            proportion = weighted_value / total_weighted_value
            component_revenue = round(original_subtotal * proportion, 2)
        else:
            # Equal split if no value data
            component_revenue = round(original_subtotal / len(pack_mappings), 2)

        # Calculate component quantity
        component_qty = original_qty * qty_multiplier

        # Calculate unit price for this component (revenue / quantity)
        component_unit_price = round(component_revenue / component_qty, 2) if component_qty > 0 else 0

        # Get additional product info from catalog
        peso_display_total = service.get_peso_display_total(component_sku)
        peso_total = service.calculate_peso_total(component_sku, component_qty)

        # Create new row for this component
        component_row = {
            **row_dict,  # Copy all original fields
            'sku': component_sku,
            'product_name': mapping['product_name'] or component_sku,
            'sku_primario': mapping['sku_primario'],
            'sku_primario_name': service.get_product_name_for_sku_primario(mapping['sku_primario']) if mapping['sku_primario'] else None,
            'category': mapping['category'] or row_dict.get('category'),
            'family': mapping['category'] or row_dict.get('family'),
            'quantity': component_qty,
            'unidades': component_qty,  # For expanded components, qty = units
            'unit_price': component_unit_price,  # Calculated price per unit
            'item_subtotal': component_revenue,
            'peso_display_total': peso_display_total,
            'peso_total': peso_total,
            'pack_parent': original_sku,  # Reference to original PACK SKU
            'is_pack_component': True,
            'pack_quantity': qty_multiplier,
            'match_type': 'pack_expansion',
            'confidence': 100,
            'in_catalog': True,
            'conversion_factor': 1
        }

        expanded_rows.append(component_row)

    return expanded_rows


def map_sku_with_quantity(sku, product_name, product_map, catalog_skus, catalog_master_skus, source=None):
    """
    Map raw SKU to catalog SKU using database-driven rules and programmatic fallbacks.

    Priority:
    1. Direct catalog match (exact SKU exists)
    2. Database mappings (sku_mappings table)
    3. Programmatic fallbacks for complex transformations

    Returns: (official_sku, match_type, pack_quantity, product_info, confidence)
    """
    import re
    from app.services.sku_mapping_service import get_sku_mapping_service

    sku = (sku or '').strip()
    product_name = product_name or ''
    source = (source or '').strip()

    # =========================================================================
    # STEP 1: Direct catalog match (highest confidence)
    # =========================================================================

    # Rule 1: Exact match in normal SKU
    if sku in catalog_skus:
        return sku, 'exact_match', 1, product_map[sku], 100

    # Rule 2: Exact match in CAJA MÁSTER
    if sku in catalog_master_skus:
        return sku, 'caja_master', 1, product_map[sku], 100

    # =========================================================================
    # STEP 2: Database-driven mappings (from sku_mappings table)
    # =========================================================================
    # This replaces hardcoded rules for: ANU- prefix, _WEB suffix, ML IDs,
    # Lokal crackers, CRSM bandeja, Keeper Pioneros, etc.

    try:
        mapping_service = get_sku_mapping_service()
        mapping_result = mapping_service.map_sku(sku, source or None)

        if mapping_result and mapping_result.target_sku:
            target_sku = mapping_result.target_sku
            # Verify target exists in product_map
            if target_sku in product_map:
                return (
                    target_sku,
                    f'db_{mapping_result.match_type}',
                    mapping_result.quantity_multiplier,
                    product_map[target_sku],
                    mapping_result.confidence
                )
            elif target_sku in catalog_master_skus and target_sku in product_map:
                return (
                    target_sku,
                    f'db_{mapping_result.match_type}+caja_master',
                    mapping_result.quantity_multiplier,
                    product_map[target_sku],
                    mapping_result.confidence
                )
    except Exception as e:
        # Log but don't fail - fall back to programmatic rules
        print(f"Warning: SKU mapping service error for '{sku}': {e}")

    # =========================================================================
    # STEP 3: Programmatic fallbacks for complex transformations
    # These rules require logic that can't easily be expressed in database
    # =========================================================================

    # NOTE: PACK prefix logic REMOVED (2025-12-18)
    # PACK products must be explicitly defined in sku_mappings table.
    # Reasons:
    # - Some PACK products are quantity multipliers (PACKGRCA_U26010 = 4× GRCA_U26010)
    # - Some PACK products are variety packs with multiple different products
    # - Extracting quantity from product_name is unreliable
    # - All PACK mappings should be explicit in the database
    # If a PACK SKU is not in sku_mappings, it should remain unmapped.

    # Trailing "20" → "10" pattern (90% confidence)
    # (Kept programmatic - regex-like transformation)
    if sku.endswith('20') and not sku.endswith('010') and not sku.endswith('020'):
        clean_sku = sku[:-2] + '10'
        if clean_sku in catalog_skus:
            return clean_sku, 'trailing_20_to_10', 1, product_map[clean_sku], 90
        if clean_sku in catalog_master_skus:
            return clean_sku, 'trailing_20_to_10+caja_master', 1, product_map[clean_sku], 90

    # Extra digits pattern (e.g., BABE_C028220 → BABE_C02810)
    # (Kept programmatic - complex regex transformation)
    extra_digit_match = re.match(r'^(.+)([0-9]{3})([0-9]{2})0$', sku)
    if extra_digit_match:
        clean_sku = extra_digit_match.group(1) + extra_digit_match.group(2) + '10'
        if clean_sku in catalog_skus:
            return clean_sku, 'extra_digits_removed', 1, product_map[clean_sku], 90
        if clean_sku in catalog_master_skus:
            return clean_sku, 'extra_digits_removed+caja_master', 1, product_map[clean_sku], 90

    # Cracker "1UES" variants (e.g., CRAA1UES → CRAA_U13510)
    # (Kept programmatic - generates target SKU dynamically)
    cracker_match = re.match(r'^CR([A-Z]{2})1UES$', sku)
    if cracker_match:
        base_code = cracker_match.group(1)
        suggested_sku = f'CR{base_code}_U13510'
        if suggested_sku in catalog_skus:
            return suggested_sku, 'cracker_1ues_variant', 1, product_map[suggested_sku], 90
        if suggested_sku in catalog_master_skus:
            return suggested_sku, 'cracker_1ues_variant+caja_master', 1, product_map[suggested_sku], 90

    # NOTE: Rules 8b (Lokal crackers), 9 (CRSM_U1000), 10 (KEEPER_PIONEROS),
    # and 13 (MercadoLibre IDs) are now handled by database mappings in STEP 2

    # Language variant suffix (_C02010 → _C02020 for English variants)
    # (Kept programmatic - needs substring replacement)
    if '_C02010' in sku:
        variant_sku = sku.replace('_C02010', '_C02020')
        if variant_sku in catalog_skus:
            return variant_sku, 'language_variant_c02010_to_c02020', 1, product_map[variant_sku], 90
        if variant_sku in catalog_master_skus:
            return variant_sku, 'language_variant_c02010_to_c02020+caja_master', 1, product_map[variant_sku], 90

    # General substring matching (85% confidence)
    # Catches any catalog SKU that appears as substring in the raw SKU
    # (Kept programmatic - dynamic search through all catalog SKUs)
    for catalog_sku in sorted(catalog_skus, key=len, reverse=True):
        if len(catalog_sku) >= 8 and catalog_sku in sku:
            return catalog_sku, 'substring_match_unitary', 1, product_map[catalog_sku], 85

    for catalog_master_sku in sorted(catalog_master_skus, key=len, reverse=True):
        if len(catalog_master_sku) >= 8 and catalog_master_sku in sku:
            return catalog_master_sku, 'substring_match_master', 1, product_map[catalog_master_sku], 85

    # No match found
    return None, 'no_match', 1, None, 0


@router.get("/data")
async def get_audit_data(
    source: Optional[List[str]] = Query(None, description="Filter by source (relbase, shopify, mercadolibre) - supports multiple"),
    category: Optional[List[str]] = Query(None, description="Filter by product category/familia (BARRAS, CRACKERS, GRANOLAS, KEEPERS) - supports multiple"),
    channel: Optional[List[str]] = Query(None, description="Filter by channel name - supports multiple"),
    customer: Optional[List[str]] = Query(None, description="Filter by customer name - supports multiple"),
    sku: Optional[str] = Query(None, description="Search in multiple fields: Cliente, Producto, Pedido, Canal, Formato, SKU Original, SKU Primario"),
    sku_primario: Optional[List[str]] = Query(None, description="Filter by SKU Primario (mapped database column) - supports multiple"),
    has_nulls: Optional[bool] = Query(None, description="Show only records with NULL values"),
    not_in_catalog: Optional[bool] = Query(None, description="Show only SKUs not in catalog"),
    from_date: Optional[str] = Query(None, description="Start date for filtering (YYYY-MM-DD format)"),
    to_date: Optional[str] = Query(None, description="End date for filtering (YYYY-MM-DD format)"),
    group_by: Optional[str] = Query(None, description="Server-side aggregation: customer_name, sku_primario, category, channel_name, order_date"),
    limit: int = Query(100, ge=1, le=1000, description="Max rows/groups to return"),
    offset: int = Query(0, ge=0, description="Offset for pagination")
):
    """
    Get comprehensive audit data with all fields for validation.

    Returns order items with:
    - Order info (ID, date, total, source)
    - Customer info (ID, name, RUT)
    - Channel info (with priority: assigned > RelBase > NULL)
    - Product info (SKU, name, category, family, format)
    - Validation flags (NULL checks, catalog checks)

    The 'sku' parameter now searches across multiple fields:
    - Customer name (Cliente)
    - Product name (Producto)
    - Order number (Pedido)
    - Channel name (Canal)
    - Product SKU (SKU Original)
    - Format and SKU Primario (filtered post-enrichment)
    """

    if not DATABASE_URL:
        raise HTTPException(status_code=500, detail="DATABASE_URL not configured")

    # Load product catalog with CAJA MÁSTER support
    product_catalog, catalog_skus, catalog_master_skus, conversion_map = load_product_mapping()

    try:
        with psycopg2.connect(DATABASE_URL) as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cursor:

                # Build WHERE clause based on filters
                where_clauses = []
                params = []

                # ✅ BASE FILTER: Only show RelBase data to avoid duplication
                where_clauses.append("o.source = 'relbase'")

                # ✅ INVOICE STATUS FILTER: Only show accepted invoices (exclude cancelled, declined, NULL)
                # This aligns with Sales Analytics endpoint and ensures data consistency
                where_clauses.append("o.invoice_status IN ('accepted', 'accepted_objection')")

                # NOTE: ANU- SKUs are now included - map_sku_with_quantity() handles mapping

                # Multi-value filters using IN operator
                if source:
                    placeholders = ','.join(['%s'] * len(source))
                    where_clauses.append(f"o.source IN ({placeholders})")
                    params.extend(source)

                # Category filter: Use CSV as source of truth instead of products table
                # This ensures consistency between filtering and enrichment
                # IMPORTANT: We need to get ALL unique SKUs from DB first, map them,
                # then filter by category to handle SKU variants (ANU- prefix, _WEB suffix, etc.)
                category_filter_skus = None
                if category:
                    # Step 1: Get all unique SKUs from DB
                    cursor.execute("""
                        SELECT DISTINCT oi.product_sku, oi.product_name, o.source
                        FROM orders o
                        JOIN order_items oi ON oi.order_id = o.id
                        WHERE oi.product_sku IS NOT NULL
                          AND oi.product_sku != ''
                    """)
                    all_db_skus = cursor.fetchall()

                    # Step 2: Map each SKU and filter by category
                    category_filter_skus = set()
                    for row in all_db_skus:
                        db_sku = row['product_sku']
                        product_name = row['product_name'] or ''
                        sku_source = row.get('source', '')

                        # Map SKU using same logic as enrichment
                        official_sku, _, _, product_info, _ = map_sku_with_quantity(
                            db_sku, product_name, product_catalog, catalog_skus, catalog_master_skus, sku_source
                        )

                        # Check if mapped SKU belongs to selected categories
                        if official_sku and product_info:
                            sku_category = product_info.get('category', '')
                            if sku_category in category:
                                category_filter_skus.add(db_sku)  # Add DB SKU (not official SKU)

                    # Step 3: Apply filter
                    if category_filter_skus:
                        placeholders = ','.join(['%s'] * len(category_filter_skus))
                        where_clauses.append(f"oi.product_sku IN ({placeholders})")
                        params.extend(list(category_filter_skus))
                    else:
                        # No SKUs found for these categories, return empty results
                        where_clauses.append("1=0")  # Always false condition

                if channel:
                    # For channel, use OR with ILIKE for partial matching
                    channel_conditions = ' OR '.join(['ch.name ILIKE %s'] * len(channel))
                    where_clauses.append(f"({channel_conditions})")
                    for ch in channel:
                        params.append(f"%{ch}%")

                if customer:
                    # For customer, use OR with ILIKE for partial matching
                    customer_conditions = ' OR '.join(['(cust_direct.name ILIKE %s OR cust_channel.name ILIKE %s)'] * len(customer))
                    where_clauses.append(f"({customer_conditions})")
                    for cust in customer:
                        params.append(f"%{cust}%")
                        params.append(f"%{cust}%")

                # SKU Primario filter (now using database column)
                if sku_primario:
                    placeholders = ','.join(['%s'] * len(sku_primario))
                    where_clauses.append(f"oi.sku_primario IN ({placeholders})")
                    params.extend(sku_primario)

                # Multi-field search: Cliente, Producto, Pedido, Canal, SKU Original
                if sku:
                    search_conditions = [
                        "COALESCE(cust_direct.name, cust_channel.name, '') ILIKE %s",  # Cliente
                        "oi.product_name ILIKE %s",                                     # Producto
                        "o.external_id ILIKE %s",                                       # Pedido
                        "ch.name ILIKE %s",                                             # Canal
                        "oi.product_sku ILIKE %s"                                       # SKU Original
                    ]
                    where_clauses.append(f"({' OR '.join(search_conditions)})")
                    # Add the same search term for each condition
                    for _ in range(len(search_conditions)):
                        params.append(f"%{sku}%")

                # Add has_nulls filter to SQL query
                if has_nulls:
                    where_clauses.append("""(
                        (cust_direct.id IS NULL AND cust_channel.id IS NULL) OR
                        o.channel_id IS NULL OR
                        oi.product_sku IS NULL OR
                        oi.product_sku = ''
                    )""")

                # Add date range filters
                if from_date and to_date:
                    where_clauses.append("o.order_date >= %s AND o.order_date <= %s")
                    params.append(from_date)
                    params.append(to_date)
                elif from_date:
                    where_clauses.append("o.order_date >= %s")
                    params.append(from_date)
                elif to_date:
                    where_clauses.append("o.order_date <= %s")
                    params.append(to_date)
                # No else clause - show ALL data when no date filter applied

                where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

                # ===== SERVER-SIDE AGGREGATION MODE =====
                # When group_by is provided, return pre-aggregated groups instead of detail items
                # This ensures group totals are accurate across ALL data, not just current page

                # ===== SPECIAL CASE: SKU PRIMARIO AGGREGATION WITH CSV MAPPING =====
                # SKU Primario requires CSV-based mapping which can't be done in SQL
                # We fetch all filtered items, apply mapping, then aggregate in Python
                if group_by == 'sku_primario':
                    # Fetch all filtered items (reasonable limit to prevent memory issues)
                    fetch_query = f"""
                        SELECT
                            o.external_id as order_external_id,
                            oi.product_sku,
                            oi.product_name,
                            o.source as order_source,
                            oi.quantity,
                            oi.subtotal
                        FROM orders o
                        LEFT JOIN order_items oi ON oi.order_id = o.id
                        LEFT JOIN products p ON p.sku = oi.product_sku
                        LEFT JOIN channels ch ON ch.id = o.channel_id
                        LEFT JOIN customers cust_direct
                            ON cust_direct.id = o.customer_id
                            AND cust_direct.source = o.source
                        LEFT JOIN LATERAL (
                            SELECT customer_external_id
                            FROM customer_channel_rules ccr
                            WHERE ccr.channel_external_id::text = (
                                CASE
                                    WHEN o.customer_notes ~ '^\\s*\\{{'
                                    THEN o.customer_notes::json->>'channel_id_relbase'
                                    ELSE NULL
                                END
                            )
                            AND ccr.is_active = TRUE
                            LIMIT 1
                        ) ccr_match ON true
                        LEFT JOIN customers cust_channel
                            ON cust_channel.external_id = ccr_match.customer_external_id
                            AND cust_channel.source = 'relbase'
                        WHERE {where_sql}
                        LIMIT 50000
                    """

                    cursor.execute(fetch_query, params)
                    all_items = cursor.fetchall()

                    # Apply CSV mapping and group in Python
                    from collections import defaultdict
                    groups = defaultdict(lambda: {
                        'pedidos': set(),
                        'cantidad': 0,
                        'total_unidades': 0,  # Add total_unidades accumulator
                        'total_revenue': 0
                    })

                    # Get ProductCatalogService for unit conversion
                    service = get_product_catalog_service()

                    for item in all_items:
                        sku = item['product_sku']
                        product_name = item['product_name']
                        order_source = item['order_source']
                        quantity = item['quantity'] or 0

                        # Map SKU to primario using same logic as enrichment
                        official_sku, _, _, _, _ = map_sku_with_quantity(
                            sku, product_name, product_catalog, catalog_skus, catalog_master_skus, order_source
                        )

                        # Get sku_primario
                        if official_sku:
                            sku_primario = get_sku_primario(official_sku, conversion_map)
                        else:
                            sku_primario = None

                        # Use 'SIN CLASIFICAR' for unmapped SKUs
                        group_key = sku_primario if sku_primario else 'SIN CLASIFICAR'

                        # Calculate converted units for this item
                        # Pass order_source to apply sku_mappings quantity_multiplier
                        converted_units = service.calculate_units(sku, quantity, order_source)

                        # Aggregate
                        groups[group_key]['pedidos'].add(item['order_external_id'])
                        groups[group_key]['cantidad'] += quantity
                        groups[group_key]['total_unidades'] += converted_units  # Accumulate converted units
                        groups[group_key]['total_revenue'] += float(item['subtotal'] or 0)

                    # Convert to list format
                    aggregated_data = []
                    for group_value, stats in groups.items():
                        # Get product name for this SKU Primario
                        sku_primario_name = service.get_product_name_for_sku_primario(group_value) if group_value != 'SIN CLASIFICAR' else None

                        aggregated_data.append({
                            'group_value': group_value,
                            'sku_primario_name': sku_primario_name,
                            'pedidos': len(stats['pedidos']),
                            'cantidad': stats['cantidad'],
                            'total_unidades': stats['total_unidades'],  # Include converted units
                            'total_revenue': stats['total_revenue']
                        })

                    # Sort by revenue descending
                    aggregated_data.sort(key=lambda x: x['total_revenue'], reverse=True)

                    # Apply pagination to groups
                    total_groups = len(aggregated_data)
                    paginated_groups = aggregated_data[offset:offset+limit]

                    # Calculate summary totals
                    total_pedidos = len(set(item['order_external_id'] for item in all_items))
                    total_quantity = sum(item['quantity'] or 0 for item in all_items)
                    total_revenue = sum(float(item['subtotal'] or 0) for item in all_items)

                    # Return aggregated response
                    return {
                        "status": "success",
                        "mode": "aggregated",
                        "group_by": "sku_primario",
                        "data": paginated_groups,
                        "summary": {
                            "total_pedidos": total_pedidos,
                            "total_unidades": total_quantity,
                            "total_revenue": total_revenue,
                            "total_groups": total_groups
                        },
                        "pagination": {
                            "limit": limit,
                            "offset": offset,
                            "total_items": total_groups,
                            "total_pages": (total_groups + limit - 1) // limit
                        }
                    }

                # Map frontend groupBy values to SQL group fields
                # sku_primario is handled above as a special case
                group_field_map = {
                    'customer_name': "COALESCE(cust_direct.name, cust_channel.name, 'SIN NOMBRE')",
                    'category': 'p.category',
                    'channel_name': "COALESCE(ch.name, 'SIN CANAL')",
                    'order_date': 'o.order_date::date',
                    'order_month': "TO_CHAR(o.order_date, 'YYYY-MM')",  # Group by year-month
                    # Additional group fields for OLAP support
                    'sku': "COALESCE(oi.product_sku, 'SIN SKU')",  # SKU Original
                    'order_external_id': 'o.external_id',  # Pedido
                    'order_source': 'o.source',  # Fuente
                    'family': "COALESCE(p.subfamily, 'SIN FAMILIA')",  # Producto
                    'format': "COALESCE(p.format, 'SIN FORMATO')"  # Formato
                }

                if group_by and group_by in group_field_map:
                    # === AGGREGATED MODE ===
                    group_field = group_field_map[group_by]

                    # Aggregation query: GROUP BY first, then paginate
                    agg_query = f"""
                        SELECT
                            {group_field} as group_value,
                            COUNT(DISTINCT o.external_id) as pedidos,
                            SUM(oi.quantity) as cantidad,
                            SUM(oi.subtotal) as total_revenue
                        FROM orders o
                        LEFT JOIN order_items oi ON oi.order_id = o.id
                        LEFT JOIN products p ON p.sku = oi.product_sku
                        LEFT JOIN channels ch ON ch.id = o.channel_id
                        LEFT JOIN customers cust_direct
                            ON cust_direct.id = o.customer_id
                            AND cust_direct.source = o.source
                        LEFT JOIN LATERAL (
                            SELECT customer_external_id
                            FROM customer_channel_rules ccr
                            WHERE ccr.channel_external_id::text = (
                                CASE
                                    WHEN o.customer_notes ~ '^\\s*\\{{'
                                    THEN o.customer_notes::json->>'channel_id_relbase'
                                    ELSE NULL
                                END
                            )
                            AND ccr.is_active = TRUE
                            LIMIT 1
                        ) ccr_match ON true
                        LEFT JOIN customers cust_channel
                            ON cust_channel.external_id = ccr_match.customer_external_id
                            AND cust_channel.source = 'relbase'
                        WHERE {where_sql}
                        GROUP BY {group_field}
                        HAVING {group_field} IS NOT NULL
                        ORDER BY total_revenue DESC
                        LIMIT %s OFFSET %s
                    """

                    # Execute aggregation query
                    params.extend([limit, offset])
                    cursor.execute(agg_query, params)
                    agg_rows = cursor.fetchall()

                    # Count total groups (not items!)
                    count_query = f"""
                        SELECT COUNT(*) as count FROM (
                            SELECT {group_field}
                            FROM orders o
                            LEFT JOIN order_items oi ON oi.order_id = o.id
                            LEFT JOIN products p ON p.sku = oi.product_sku
                            LEFT JOIN channels ch ON ch.id = o.channel_id
                            LEFT JOIN customers cust_direct
                                ON cust_direct.id = o.customer_id
                                AND cust_direct.source = o.source
                            LEFT JOIN LATERAL (
                                SELECT customer_external_id
                                FROM customer_channel_rules ccr
                                WHERE ccr.channel_external_id::text = (
                                    CASE
                                        WHEN o.customer_notes ~ '^\\s*\\{{'
                                        THEN o.customer_notes::json->>'channel_id_relbase'
                                        ELSE NULL
                                    END
                                )
                                AND ccr.is_active = TRUE
                                LIMIT 1
                            ) ccr_match ON true
                            LEFT JOIN customers cust_channel
                                ON cust_channel.external_id = ccr_match.customer_external_id
                                AND cust_channel.source = 'relbase'
                            WHERE {where_sql}
                            GROUP BY {group_field}
                            HAVING {group_field} IS NOT NULL
                        ) subquery
                    """

                    cursor.execute(count_query, params[:-2])  # Exclude limit/offset
                    total_groups = cursor.fetchone()['count']

                    # Summary totals (same across ALL filtered data)
                    summary_query = f"""
                        SELECT
                            COUNT(DISTINCT o.external_id) as total_pedidos,
                            SUM(oi.quantity) as total_quantity,
                            SUM(oi.subtotal) as total_revenue
                        FROM orders o
                        LEFT JOIN order_items oi ON oi.order_id = o.id
                        LEFT JOIN products p ON p.sku = oi.product_sku
                        LEFT JOIN channels ch ON ch.id = o.channel_id
                        LEFT JOIN customers cust_direct
                            ON cust_direct.id = o.customer_id
                            AND cust_direct.source = o.source
                        LEFT JOIN LATERAL (
                            SELECT customer_external_id
                            FROM customer_channel_rules ccr
                            WHERE ccr.channel_external_id::text = (
                                CASE
                                    WHEN o.customer_notes ~ '^\\s*\\{{'
                                    THEN o.customer_notes::json->>'channel_id_relbase'
                                    ELSE NULL
                                END
                            )
                            AND ccr.is_active = TRUE
                            LIMIT 1
                        ) ccr_match ON true
                        LEFT JOIN customers cust_channel
                            ON cust_channel.external_id = ccr_match.customer_external_id
                            AND cust_channel.source = 'relbase'
                        WHERE {where_sql}
                    """

                    cursor.execute(summary_query, params[:-2])  # Exclude limit/offset
                    summary_row = cursor.fetchone()

                    # Convert aggregated rows to dicts and calculate proper total_unidades
                    aggregated_data = []
                    for row in agg_rows:
                        row_dict = dict(row)
                        # Convert Decimal to float for JSON serialization
                        row_dict['total_revenue'] = float(row_dict['total_revenue'] or 0)

                        # Calculate actual total_unidades by fetching and converting units
                        # Query order_items for this group and sum converted units
                        group_value = row_dict['group_value']

                        # Build query to get all order_items in this group
                        unidades_query = f"""
                            SELECT
                                oi.product_sku,
                                oi.quantity,
                                o.source as order_source
                            FROM orders o
                            LEFT JOIN order_items oi ON oi.order_id = o.id
                            LEFT JOIN products p ON p.sku = oi.product_sku
                            LEFT JOIN channels ch ON ch.id = o.channel_id
                            LEFT JOIN customers cust_direct
                                ON cust_direct.id = o.customer_id
                                AND cust_direct.source = o.source
                            LEFT JOIN LATERAL (
                                SELECT customer_external_id
                                FROM customer_channel_rules ccr
                                WHERE ccr.channel_external_id::text = (
                                    CASE
                                        WHEN o.customer_notes ~ '^\\s*\\{{'
                                        THEN o.customer_notes::json->>'channel_id_relbase'
                                        ELSE NULL
                                    END
                                )
                                AND ccr.is_active = TRUE
                                LIMIT 1
                            ) ccr_match ON true
                            LEFT JOIN customers cust_channel
                                ON cust_channel.external_id = ccr_match.customer_external_id
                                AND cust_channel.source = 'relbase'
                            WHERE {where_sql}
                              AND {group_field} = %s
                              AND oi.product_sku IS NOT NULL
                        """

                        cursor.execute(unidades_query, params[:-2] + [group_value])
                        group_items = cursor.fetchall()

                        # Calculate total converted units for this group
                        total_unidades = 0
                        service = get_product_catalog_service()
                        for item in group_items:
                            sku = item['product_sku']
                            quantity = item['quantity'] or 0
                            order_source = item.get('order_source')
                            # Use ProductCatalogService to calculate converted units
                            # Pass source to apply sku_mappings quantity_multiplier
                            converted_units = service.calculate_units(sku, quantity, order_source)
                            total_unidades += converted_units

                        # Add the properly calculated total_unidades to the row
                        row_dict['total_unidades'] = total_unidades

                        aggregated_data.append(row_dict)

                    # Return aggregated response
                    return {
                        "status": "success",
                        "mode": "aggregated",
                        "group_by": group_by,
                        "data": aggregated_data,
                        "summary": {
                            "total_pedidos": summary_row['total_pedidos'] or 0,
                            "total_unidades": summary_row['total_quantity'] or 0,  # Note: approximation
                            "total_revenue": float(summary_row['total_revenue'] or 0),
                            "total_groups": total_groups
                        },
                        "pagination": {
                            "limit": limit,
                            "offset": offset,
                            "total_items": total_groups,  # Total groups, not items
                            "total_pages": (total_groups + limit - 1) // limit
                        }
                    }

                # ===== DETAIL MODE (No grouping) =====
                # Continue with existing logic for individual items

                # Pre-determine mapped SKUs if not_in_catalog filter is active
                mapped_skus_set = None
                if not_in_catalog:
                    # Get all unique SKUs with their product names and source
                    cursor.execute(f"""
                        SELECT DISTINCT oi.product_sku as sku, oi.product_name, o.source
                        FROM orders o
                        LEFT JOIN order_items oi ON oi.order_id = o.id
                        LEFT JOIN channels ch ON ch.id = o.channel_id
                        LEFT JOIN customers cust_direct
                            ON cust_direct.id = o.customer_id
                            AND cust_direct.source = o.source
                        LEFT JOIN LATERAL (
                            SELECT customer_external_id
                            FROM customer_channel_rules ccr
                            WHERE ccr.channel_external_id::text = (
                                CASE
                                    WHEN o.customer_notes ~ '^\\s*\\{{'
                                    THEN o.customer_notes::json->>'channel_id_relbase'
                                    ELSE NULL
                                END
                            )
                            AND ccr.is_active = TRUE
                            LIMIT 1
                        ) ccr_match ON true
                        LEFT JOIN customers cust_channel
                            ON cust_channel.external_id = ccr_match.customer_external_id
                            AND cust_channel.source = 'relbase'
                        WHERE {where_sql}
                          AND oi.product_sku IS NOT NULL
                          AND oi.product_sku != ''
                    """, params)
                    all_skus = cursor.fetchall()

                    # Determine which SKUs are mapped
                    mapped_skus_set = set()
                    unmapped_skus_set = set()
                    for row in all_skus:
                        sku_val = row['sku']
                        product_name = row['product_name'] or ''
                        sku_source = row.get('source', '')
                        official_sku, _, _, _, _ = map_sku_with_quantity(
                            sku_val, product_name, product_catalog, catalog_skus, catalog_master_skus, sku_source
                        )
                        if official_sku:
                            mapped_skus_set.add(sku_val)
                        else:
                            unmapped_skus_set.add(sku_val)

                    # Add filter to WHERE clause for unmapped SKUs only
                    if unmapped_skus_set:
                        # Use SQL IN clause for unmapped SKUs
                        placeholders = ','.join(['%s'] * len(unmapped_skus_set))
                        where_clauses.append(f"oi.product_sku IN ({placeholders})")
                        params.extend(list(unmapped_skus_set))
                        where_sql = " AND ".join(where_clauses)
                    else:
                        # No unmapped SKUs, return empty result
                        return {
                            "status": "success",
                            "data": [],
                            "meta": {
                                "total": 0,
                                "limit": limit,
                                "offset": offset,
                                "returned": 0,
                                "filters": {
                                    "source": source,
                                    "channel": channel,
                                    "customer": customer,
                                    "sku": sku,
                                    "has_nulls": has_nulls,
                                    "not_in_catalog": not_in_catalog,
                                    "from_date": from_date,
                                    "to_date": to_date
                                }
                            }
                        }

                # Main query with filters applied in SQL
                query = f"""
                    SELECT
                        -- Order info
                        o.id as order_id,
                        o.external_id as order_external_id,
                        o.order_date,
                        o.total as order_total,
                        o.source as order_source,

                        -- Customer info (improved with channel-based mapping)
                        COALESCE(
                            cust_direct.id::text,
                            cust_channel.id::text,
                            'NULL'
                        ) as customer_id,
                        COALESCE(
                            cust_direct.external_id,
                            cust_channel.external_id,
                            'NULL'
                        ) as customer_external_id,
                        COALESCE(
                            cust_direct.name,
                            cust_channel.name,
                            'SIN NOMBRE'
                        ) as customer_name,
                        CASE
                            WHEN cust_direct.id IS NOT NULL THEN 'direct'
                            WHEN cust_channel.id IS NOT NULL THEN 'channel_mapped'
                            ELSE 'null'
                        END as customer_source,

                        -- Channel info (enriched with channel names from channels table)
                        COALESCE(
                            ch.name,
                            'SIN CANAL'
                        ) as channel_name,
                        COALESCE(
                            o.channel_id::text,
                            'NULL'
                        ) as channel_id,
                        CASE
                            WHEN ch.name IS NOT NULL THEN 'relbase'
                            ELSE 'null'
                        END as channel_source,

                        -- Order Item info
                        oi.id as item_id,
                        oi.product_sku as sku,
                        oi.product_name,
                        oi.quantity,
                        ROUND(oi.subtotal / NULLIF(oi.quantity, 0), 2) as unit_price,  -- Real net unit price (after discounts, without IVA)
                        oi.subtotal as item_subtotal,

                        -- Product Mapping (will be enriched with CSV data)
                        p.category,
                        p.subfamily as family,
                        p.format,

                        -- Flags
                        CASE WHEN cust_direct.id IS NULL AND cust_channel.id IS NULL THEN true ELSE false END as customer_null,
                        CASE WHEN o.channel_id IS NULL THEN true ELSE false END as channel_null,
                        CASE WHEN oi.product_sku IS NULL OR oi.product_sku = '' THEN true ELSE false END as sku_null

                    FROM orders o
                    LEFT JOIN order_items oi ON oi.order_id = o.id
                    LEFT JOIN products p ON p.sku = oi.product_sku
                    LEFT JOIN channels ch ON ch.id = o.channel_id

                    -- Customer Joins (Direct + Channel-based mapping)
                    LEFT JOIN customers cust_direct
                        ON cust_direct.id = o.customer_id
                        AND cust_direct.source = o.source

                    -- Channel-based customer mapping via customer_channel_rules
                    LEFT JOIN LATERAL (
                        SELECT customer_external_id
                        FROM customer_channel_rules ccr
                        WHERE ccr.channel_external_id::text = (
                            CASE
                                WHEN o.customer_notes ~ '^\\s*\\{{'
                                THEN o.customer_notes::json->>'channel_id_relbase'
                                ELSE NULL
                            END
                        )
                        AND ccr.is_active = TRUE
                        LIMIT 1
                    ) ccr_match ON true

                    LEFT JOIN customers cust_channel
                        ON cust_channel.external_id = ccr_match.customer_external_id
                        AND cust_channel.source = 'relbase'

                    WHERE {where_sql}
                    ORDER BY o.order_date DESC, o.id, oi.id
                    LIMIT %s OFFSET %s
                """

                params.extend([limit, offset])
                cursor.execute(query, params)
                rows = cursor.fetchall()

                # Get total count for pagination with filters applied
                count_query = f"""
                    SELECT COUNT(DISTINCT oi.id)
                    FROM orders o
                    LEFT JOIN order_items oi ON oi.order_id = o.id
                    LEFT JOIN products p ON p.sku = oi.product_sku
                    LEFT JOIN channels ch ON ch.id = o.channel_id
                    LEFT JOIN customers cust_direct
                        ON cust_direct.id = o.customer_id
                        AND cust_direct.source = o.source
                    LEFT JOIN LATERAL (
                        SELECT customer_external_id
                        FROM customer_channel_rules ccr
                        WHERE ccr.channel_external_id::text = (
                            CASE
                                WHEN o.customer_notes ~ '^\\s*\\{{'
                                THEN o.customer_notes::json->>'channel_id_relbase'
                                ELSE NULL
                            END
                        )
                        AND ccr.is_active = TRUE
                        LIMIT 1
                    ) ccr_match ON true
                    LEFT JOIN customers cust_channel
                        ON cust_channel.external_id = ccr_match.customer_external_id
                        AND cust_channel.source = 'relbase'
                    WHERE {where_sql}
                """

                cursor.execute(count_query, params[:-2])  # Exclude limit/offset
                total_count = cursor.fetchone()['count']

                # ===== SUMMARY TOTALS QUERY =====
                # Calculate totals for ALL filtered data (not just current page)
                # This provides accurate totals for summary cards regardless of pagination
                summary_query = f"""
                    SELECT
                        COUNT(DISTINCT o.external_id) as total_pedidos,
                        SUM(oi.quantity) as total_quantity,
                        SUM(oi.subtotal) as total_revenue
                    FROM orders o
                    LEFT JOIN order_items oi ON oi.order_id = o.id
                    LEFT JOIN products p ON p.sku = oi.product_sku
                    LEFT JOIN channels ch ON ch.id = o.channel_id
                    LEFT JOIN customers cust_direct
                        ON cust_direct.id = o.customer_id
                        AND cust_direct.source = o.source
                    LEFT JOIN LATERAL (
                        SELECT customer_external_id
                        FROM customer_channel_rules ccr
                        WHERE ccr.channel_external_id::text = (
                            CASE
                                WHEN o.customer_notes ~ '^\\s*\\{{'
                                THEN o.customer_notes::json->>'channel_id_relbase'
                                ELSE NULL
                            END
                        )
                        AND ccr.is_active = TRUE
                        LIMIT 1
                    ) ccr_match ON true
                    LEFT JOIN customers cust_channel
                        ON cust_channel.external_id = ccr_match.customer_external_id
                        AND cust_channel.source = 'relbase'
                    WHERE {where_sql}
                """

                cursor.execute(summary_query, params[:-2])  # Exclude limit/offset
                summary_row = cursor.fetchone()

                # Extract summary values (will be refined with product mapping)
                filtered_totals = {
                    "total_pedidos": summary_row['total_pedidos'] or 0,
                    "total_quantity": summary_row['total_quantity'] or 0,  # Raw quantity (will adjust for unidades)
                    "total_revenue": float(summary_row['total_revenue'] or 0)
                }

                # ===== UNITS TOTALS FROM MV =====
                # Get accurate unit totals from sales_facts_mv (applies conversion factors)
                # This ensures consistency with Sales Analytics view
                mv_units_query = """
                    SELECT
                        SUM(
                            CASE
                                WHEN is_caja_master THEN units_sold * COALESCE(items_per_master_box, 1)
                                ELSE units_sold * COALESCE(units_per_display, 1)
                            END
                        ) as total_units
                    FROM sales_facts_mv
                    WHERE order_date >= %s AND order_date <= %s
                """
                # Get date params - dates are added LAST to params (before limit/offset)
                # So we need to extract from the end, not the beginning
                date_params = []
                if from_date and to_date:
                    date_params = [from_date, to_date]
                elif from_date:
                    date_params = [from_date, from_date]  # Use same date for range
                elif to_date:
                    date_params = [to_date, to_date]  # Use same date for range

                if len(date_params) >= 2:
                    cursor.execute(mv_units_query, date_params[:2])
                    mv_row = cursor.fetchone()
                    filtered_totals["total_unidades_mv"] = int(mv_row['total_units'] or 0) if mv_row else 0
                else:
                    filtered_totals["total_unidades_mv"] = 0

                # Enrich with product catalog data using conservative mapping
                enriched_rows = []
                for row in rows:
                    row_dict = dict(row)
                    sku = row_dict.get('sku', '')
                    product_name = row_dict.get('product_name', '')
                    order_source = row_dict.get('order_source', '')
                    quantity = row_dict.get('quantity', 0)

                    # Apply conservative mapping with quantity extraction
                    official_sku, match_type, pack_qty, product_info, confidence = map_sku_with_quantity(
                        sku, product_name, product_catalog, catalog_skus, catalog_master_skus, order_source
                    )

                    if official_sku:
                        # Product successfully mapped
                        # Get SKU Primario
                        sku_primario = get_sku_primario(official_sku, conversion_map)

                        # Get SKU Primario Name (formatted)
                        service = get_product_catalog_service()
                        sku_primario_name = service.get_product_name_for_sku_primario(sku_primario) if sku_primario else None

                        # Calculate total units
                        # IMPORTANT: Pass original SKU (not official_sku) to check sku_mappings for multipliers
                        # e.g., KEEPERPACK (original) → finds mapping ×5 → 7 × 5 = 35 units
                        unidades = calculate_units(sku, quantity, conversion_map, order_source)

                        # Calculate peso (weight) - Phase 4
                        peso_display_total = service.get_peso_display_total(official_sku)
                        peso_total = service.calculate_peso_total(official_sku, quantity)

                        row_dict['sku_primario'] = sku_primario
                        row_dict['sku_primario_name'] = sku_primario_name
                        row_dict['unidades'] = unidades
                        row_dict['peso_display_total'] = peso_display_total
                        row_dict['peso_total'] = peso_total
                        row_dict['pack_quantity'] = pack_qty
                        row_dict['match_type'] = match_type
                        row_dict['confidence'] = confidence
                        row_dict['category'] = product_info['category'] or row_dict.get('category')
                        row_dict['family'] = product_info['family'] or row_dict.get('family')
                        row_dict['format'] = product_info['format'] or row_dict.get('format')
                        row_dict['in_catalog'] = True

                        # Derive effective conversion factor from actual calculation
                        # This ensures the displayed multiplier matches the units calculation
                        # Important: Uses unidades/quantity to capture sku_mappings multipliers
                        # e.g., GRAL_C02010: 13 qty → 130 units → shows (x10)
                        row_dict['conversion_factor'] = unidades // quantity if quantity > 0 else 1
                    else:
                        # Not mapped
                        row_dict['sku_primario'] = None
                        row_dict['sku_primario_name'] = None
                        row_dict['unidades'] = quantity  # Default to quantity if not mapped
                        row_dict['peso_display_total'] = 0.0
                        row_dict['peso_total'] = 0.0
                        row_dict['pack_quantity'] = 1
                        row_dict['match_type'] = 'no_match'
                        row_dict['confidence'] = 0
                        row_dict['in_catalog'] = False
                        row_dict['conversion_factor'] = 1

                    # ✅ PACK EXPANSION: Check if this is a variety pack (multiple component mappings)
                    # If so, expand into individual component rows with proportional revenue
                    pack_mappings = get_pack_component_mappings(sku, cursor)
                    if pack_mappings:
                        # Expand PACK into component rows
                        component_rows = expand_pack_to_components(row_dict, pack_mappings, cursor)
                        enriched_rows.extend(component_rows)
                    else:
                        # Normal product - add as-is
                        enriched_rows.append(row_dict)

                # Calculate totals from enriched data
                # total_unidades: Use MV-based calculation for accuracy (full dataset)
                # Fallback to page sum if MV returns 0 (no date filter case)
                # total_peso: Calculated from current page only (acceptable approximation)
                page_unidades = sum(row.get('unidades', 0) for row in enriched_rows)
                mv_unidades = filtered_totals.get("total_unidades_mv", 0)
                filtered_totals["total_unidades"] = mv_unidades if mv_unidades > 0 else page_unidades
                filtered_totals["total_peso"] = round(sum(row.get('peso_total', 0) for row in enriched_rows), 4)
                # Also keep page-only units for debugging
                filtered_totals["page_unidades"] = page_unidades

                return {
                    "status": "success",
                    "data": enriched_rows,
                    "meta": {
                        "total": total_count,
                        "limit": limit,
                        "offset": offset,
                        "returned": len(enriched_rows),
                        "filters": {
                            "source": source,
                            "category": category,
                            "channel": channel,
                            "customer": customer,
                            "sku": sku,
                            "has_nulls": has_nulls,
                            "not_in_catalog": not_in_catalog,
                            "from_date": from_date,
                            "to_date": to_date
                        }
                    },
                    "summary": {
                        "total_pedidos": filtered_totals["total_pedidos"],
                        "total_unidades": filtered_totals["total_unidades"],
                        "total_peso": filtered_totals["total_peso"],
                        "total_revenue": filtered_totals["total_revenue"]
                    }
                }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching audit data: {str(e)}")


@router.get("/filters")
async def get_available_filters():
    """
    Get available filter values (unique values for dropdowns).
    Returns unique sources, channels, customers, SKUs for filtering.
    """

    if not DATABASE_URL:
        raise HTTPException(status_code=500, detail="DATABASE_URL not configured")

    try:
        with psycopg2.connect(DATABASE_URL) as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cursor:

                # Get unique sources
                cursor.execute("""
                    SELECT DISTINCT source
                    FROM orders
                    WHERE EXTRACT(YEAR FROM order_date) = 2025
                    ORDER BY source
                """)
                sources = [row['source'] for row in cursor.fetchall() if row['source']]

                # Get official Relbase channels that have data for the current year
                # Exclude channels with no orders (EXPORTACIÓN, HORECA, MARKETPLACES)
                # Exclude "Relbase" as it's a data source, not a sales channel
                current_year = datetime.now().year
                cursor.execute("""
                    SELECT DISTINCT ch.name as channel_name
                    FROM channels ch
                    INNER JOIN orders o ON o.channel_id = ch.id
                    WHERE ch.external_id IS NOT NULL
                      AND ch.source = 'relbase'
                      AND ch.is_active = true
                      AND ch.name NOT ILIKE 'relbase'
                      AND o.source = 'relbase'
                      AND o.invoice_status IN ('accepted', 'accepted_objection')
                      AND EXTRACT(YEAR FROM o.order_date) = %s
                    ORDER BY ch.name
                """, (current_year,))
                channels = [row['channel_name'] for row in cursor.fetchall()]

                # Always add "Sin Canal Asignado" as default for orders without channel
                channels.append('Sin Canal Asignado')

                # Get unique customers (limit to top 100 by order count)
                cursor.execute("""
                    SELECT cust.name as customer_name
                    FROM orders o
                    LEFT JOIN customers cust ON cust.id = o.customer_id
                    WHERE EXTRACT(YEAR FROM order_date) = %s
                      AND cust.name IS NOT NULL
                    GROUP BY cust.name
                    ORDER BY COUNT(*) DESC
                    LIMIT 100
                """, (current_year,))
                customers = [row['customer_name'] for row in cursor.fetchall()]

                # Get unique SKUs (limit to top 100 by quantity sold)
                cursor.execute("""
                    SELECT oi.product_sku as sku
                    FROM order_items oi
                    JOIN orders o ON o.id = oi.order_id
                    WHERE EXTRACT(YEAR FROM o.order_date) = %s
                      AND oi.product_sku IS NOT NULL
                      AND oi.product_sku != ''
                    GROUP BY oi.product_sku
                    ORDER BY SUM(oi.quantity) DESC
                    LIMIT 100
                """, (current_year,))
                skus = [row['sku'] for row in cursor.fetchall()]

                return {
                    "status": "success",
                    "data": {
                        "sources": sources,
                        "channels": channels,
                        "customers": customers,
                        "skus": skus
                    }
                }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching filters: {str(e)}")


@router.get("/summary")
async def get_audit_summary():
    """
    Get audit summary statistics.
    Returns counts of NULL values, unmapped products, data quality metrics.
    Uses conservative mapping strategy to calculate actual coverage.
    """

    if not DATABASE_URL:
        raise HTTPException(status_code=500, detail="DATABASE_URL not configured")

    # Load product catalog with CAJA MÁSTER support
    product_catalog, catalog_skus, catalog_master_skus, conversion_map = load_product_mapping()

    try:
        with psycopg2.connect(DATABASE_URL) as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cursor:

                # Total orders
                cursor.execute("""
                    SELECT COUNT(*) as total
                    FROM orders
                    WHERE EXTRACT(YEAR FROM order_date) = 2025
                """)
                total_orders = cursor.fetchone()['total']

                # Orders with NULL customer
                cursor.execute("""
                    SELECT COUNT(*) as total
                    FROM orders
                    WHERE EXTRACT(YEAR FROM order_date) = 2025
                      AND customer_id IS NULL
                """)
                null_customers = cursor.fetchone()['total']

                # Orders with NULL channel
                cursor.execute("""
                    SELECT COUNT(*) as total
                    FROM orders o
                    WHERE EXTRACT(YEAR FROM order_date) = 2025
                      AND o.channel_id IS NULL
                """)
                null_channels = cursor.fetchone()['total']

                # Order items with NULL or empty SKU
                cursor.execute("""
                    SELECT COUNT(*) as total
                    FROM order_items oi
                    JOIN orders o ON o.id = oi.order_id
                    WHERE EXTRACT(YEAR FROM o.order_date) = 2025
                      AND (oi.product_sku IS NULL OR oi.product_sku = '')
                """)
                null_skus = cursor.fetchone()['total']

                # Unique SKUs
                cursor.execute("""
                    SELECT COUNT(DISTINCT oi.product_sku) as total
                    FROM order_items oi
                    JOIN orders o ON o.id = oi.order_id
                    WHERE EXTRACT(YEAR FROM o.order_date) = 2025
                      AND oi.product_sku IS NOT NULL
                      AND oi.product_sku != ''
                """)
                unique_skus = cursor.fetchone()['total']

                # Get all SKUs with product names and source for mapping
                cursor.execute("""
                    SELECT DISTINCT oi.product_sku as sku, oi.product_name, o.source
                    FROM order_items oi
                    JOIN orders o ON o.id = oi.order_id
                    WHERE EXTRACT(YEAR FROM o.order_date) = 2025
                      AND oi.product_sku IS NOT NULL
                      AND oi.product_sku != ''
                """)
                all_skus = cursor.fetchall()

                # Apply conservative mapping to determine which SKUs are actually unmapped
                unmapped_skus = []
                mapped_skus = []
                for row in all_skus:
                    sku = row['sku']
                    product_name = row['product_name'] or ''
                    sku_source = row.get('source', '')

                    official_sku, match_type, pack_qty, product_info, confidence = map_sku_with_quantity(
                        sku, product_name, product_catalog, catalog_skus, catalog_master_skus, sku_source
                    )

                    if official_sku:
                        mapped_skus.append(sku)
                    else:
                        unmapped_skus.append(sku)

                return {
                    "status": "success",
                    "data": {
                        "total_orders": total_orders,
                        "data_quality": {
                            "null_customers": null_customers,
                            "null_channels": null_channels,
                            "null_skus": null_skus,
                            "completeness_pct": round((1 - (null_customers + null_channels + null_skus) / (total_orders * 3)) * 100, 2) if total_orders > 0 else 0
                        },
                        "product_mapping": {
                            "unique_skus": unique_skus,
                            "mapped_skus": len(mapped_skus),
                            "not_in_catalog": len(unmapped_skus),
                            "catalog_coverage_pct": round((len(mapped_skus) / unique_skus) * 100, 2) if unique_skus > 0 else 0,
                            "unmapped_skus_sample": unmapped_skus[:20]  # First 20 unmapped SKUs
                        }
                    }
                }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching audit summary: {str(e)}")


@router.get("/export")
async def export_audit_data(
    source: Optional[List[str]] = Query(None, description="Filter by source"),
    category: Optional[List[str]] = Query(None, description="Filter by product category/familia"),
    channel: Optional[List[str]] = Query(None, description="Filter by channel name"),
    customer: Optional[List[str]] = Query(None, description="Filter by customer name"),
    sku: Optional[str] = Query(None, description="Search in multiple fields"),
    sku_primario: Optional[List[str]] = Query(None, description="Filter by SKU Primario"),
    has_nulls: Optional[bool] = Query(None, description="Show only records with NULL values"),
    not_in_catalog: Optional[bool] = Query(None, description="Show only SKUs not in catalog"),
    from_date: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    group_by: Optional[str] = Query(None, description="Grouping field for aggregated export")
):
    """
    Export audit data to Excel file.

    Supports both detail mode (no grouping) and aggregated mode (with group_by).
    Returns an Excel file with properly formatted data matching the current filters.
    """

    if not DATABASE_URL:
        raise HTTPException(status_code=500, detail="DATABASE_URL not configured")

    # Load product catalog
    product_catalog, catalog_skus, catalog_master_skus, conversion_map = load_product_mapping()

    try:
        with psycopg2.connect(DATABASE_URL) as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cursor:

                # Build WHERE clause (same as /data endpoint)
                where_clauses = []
                params = []

                # Base filters
                where_clauses.append("o.source = 'relbase'")
                where_clauses.append("o.invoice_status IN ('accepted', 'accepted_objection')")
                # NOTE: ANU- SKUs are now included - map_sku_with_quantity() handles mapping

                if source:
                    placeholders = ','.join(['%s'] * len(source))
                    where_clauses.append(f"o.source IN ({placeholders})")
                    params.extend(source)

                # Category filter with SKU mapping
                category_filter_skus = None
                if category:
                    cursor.execute("""
                        SELECT DISTINCT oi.product_sku, oi.product_name, o.source
                        FROM orders o
                        JOIN order_items oi ON oi.order_id = o.id
                        WHERE oi.product_sku IS NOT NULL AND oi.product_sku != ''
                    """)
                    all_db_skus = cursor.fetchall()

                    category_filter_skus = set()
                    for row in all_db_skus:
                        db_sku = row['product_sku']
                        product_name = row['product_name'] or ''
                        sku_source = row.get('source', '')
                        official_sku, _, _, product_info, _ = map_sku_with_quantity(
                            db_sku, product_name, product_catalog, catalog_skus, catalog_master_skus, sku_source
                        )
                        if official_sku and product_info:
                            sku_category = product_info.get('category', '')
                            if sku_category in category:
                                category_filter_skus.add(db_sku)

                    if category_filter_skus:
                        placeholders = ','.join(['%s'] * len(category_filter_skus))
                        where_clauses.append(f"oi.product_sku IN ({placeholders})")
                        params.extend(list(category_filter_skus))
                    else:
                        where_clauses.append("1=0")

                if channel:
                    channel_conditions = ' OR '.join(['ch.name ILIKE %s'] * len(channel))
                    where_clauses.append(f"({channel_conditions})")
                    for ch in channel:
                        params.append(f"%{ch}%")

                if customer:
                    customer_conditions = ' OR '.join(['(cust_direct.name ILIKE %s OR cust_channel.name ILIKE %s)'] * len(customer))
                    where_clauses.append(f"({customer_conditions})")
                    for cust in customer:
                        params.append(f"%{cust}%")
                        params.append(f"%{cust}%")

                if sku_primario:
                    placeholders = ','.join(['%s'] * len(sku_primario))
                    where_clauses.append(f"oi.sku_primario IN ({placeholders})")
                    params.extend(sku_primario)

                if sku:
                    search_conditions = [
                        "COALESCE(cust_direct.name, cust_channel.name, '') ILIKE %s",
                        "oi.product_name ILIKE %s",
                        "o.external_id ILIKE %s",
                        "ch.name ILIKE %s",
                        "oi.product_sku ILIKE %s"
                    ]
                    where_clauses.append(f"({' OR '.join(search_conditions)})")
                    for _ in range(len(search_conditions)):
                        params.append(f"%{sku}%")

                if has_nulls:
                    where_clauses.append("""(
                        (cust_direct.id IS NULL AND cust_channel.id IS NULL) OR
                        o.channel_id IS NULL OR
                        oi.product_sku IS NULL OR
                        oi.product_sku = ''
                    )""")

                if from_date and to_date:
                    where_clauses.append("o.order_date >= %s AND o.order_date <= %s")
                    params.append(from_date)
                    params.append(to_date)
                elif from_date:
                    where_clauses.append("o.order_date >= %s")
                    params.append(from_date)
                elif to_date:
                    where_clauses.append("o.order_date <= %s")
                    params.append(to_date)

                where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"

                # Create workbook
                wb = Workbook()
                ws = wb.active

                # Styles
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                number_format_currency = '#,##0'
                number_format_decimal = '#,##0.00'

                # ===== AGGREGATED MODE EXPORT =====
                if group_by == 'sku_primario':
                    ws.title = "Agrupado por SKU Primario"

                    # Headers
                    headers = ["SKU Primario", "Producto", "Pedidos", "Cantidad", "Unidades", "Revenue"]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border

                    # Fetch all data for aggregation
                    fetch_query = f"""
                        SELECT
                            o.external_id as order_external_id,
                            oi.product_sku,
                            oi.product_name,
                            o.source as order_source,
                            oi.quantity,
                            oi.subtotal
                        FROM orders o
                        LEFT JOIN order_items oi ON oi.order_id = o.id
                        LEFT JOIN products p ON p.sku = oi.product_sku
                        LEFT JOIN channels ch ON ch.id = o.channel_id
                        LEFT JOIN customers cust_direct ON cust_direct.id = o.customer_id AND cust_direct.source = o.source
                        LEFT JOIN LATERAL (
                            SELECT customer_external_id FROM customer_channel_rules ccr
                            WHERE ccr.channel_external_id::text = (
                                CASE WHEN o.customer_notes ~ '^\\s*\\{{'
                                THEN o.customer_notes::json->>'channel_id_relbase' ELSE NULL END
                            ) AND ccr.is_active = TRUE LIMIT 1
                        ) ccr_match ON true
                        LEFT JOIN customers cust_channel ON cust_channel.external_id = ccr_match.customer_external_id AND cust_channel.source = 'relbase'
                        WHERE {where_sql}
                        LIMIT 100000
                    """

                    cursor.execute(fetch_query, params)
                    all_items = cursor.fetchall()

                    # Aggregate in Python
                    from collections import defaultdict
                    groups = defaultdict(lambda: {'pedidos': set(), 'cantidad': 0, 'total_unidades': 0, 'total_revenue': 0})
                    service = get_product_catalog_service()

                    for item in all_items:
                        item_sku = item['product_sku']
                        product_name = item['product_name']
                        order_source = item['order_source']
                        quantity = item['quantity'] or 0

                        official_sku, _, _, _, _ = map_sku_with_quantity(
                            item_sku, product_name, product_catalog, catalog_skus, catalog_master_skus, order_source
                        )

                        sku_prim = get_sku_primario(official_sku, conversion_map) if official_sku else None
                        group_key = sku_prim if sku_prim else 'SIN CLASIFICAR'
                        converted_units = service.calculate_units(item_sku, quantity, order_source)

                        groups[group_key]['pedidos'].add(item['order_external_id'])
                        groups[group_key]['cantidad'] += quantity
                        groups[group_key]['total_unidades'] += converted_units
                        groups[group_key]['total_revenue'] += float(item['subtotal'] or 0)

                    # Sort and write data
                    sorted_groups = sorted(groups.items(), key=lambda x: x[1]['total_revenue'], reverse=True)

                    for row_idx, (group_value, stats) in enumerate(sorted_groups, 2):
                        sku_primario_name = service.get_product_name_for_sku_primario(group_value) if group_value != 'SIN CLASIFICAR' else None

                        ws.cell(row=row_idx, column=1, value=group_value).border = thin_border
                        ws.cell(row=row_idx, column=2, value=sku_primario_name or '').border = thin_border
                        ws.cell(row=row_idx, column=3, value=len(stats['pedidos'])).border = thin_border
                        ws.cell(row=row_idx, column=4, value=stats['cantidad']).border = thin_border
                        ws.cell(row=row_idx, column=5, value=stats['total_unidades']).border = thin_border
                        cell = ws.cell(row=row_idx, column=6, value=stats['total_revenue'])
                        cell.border = thin_border
                        cell.number_format = number_format_currency

                    # Column widths
                    ws.column_dimensions['A'].width = 20
                    ws.column_dimensions['B'].width = 40
                    ws.column_dimensions['C'].width = 12
                    ws.column_dimensions['D'].width = 12
                    ws.column_dimensions['E'].width = 12
                    ws.column_dimensions['F'].width = 15

                # ===== OTHER AGGREGATED MODES =====
                elif group_by and group_by in ['customer_name', 'channel_name', 'order_month', 'family', 'format', 'sku']:
                    group_field_map = {
                        'customer_name': ("COALESCE(cust_direct.name, cust_channel.name, 'SIN NOMBRE')", "Cliente"),
                        'channel_name': ("COALESCE(ch.name, 'SIN CANAL')", "Canal"),
                        'order_month': ("TO_CHAR(o.order_date, 'YYYY-MM')", "Mes"),
                        'family': ("COALESCE(p.subfamily, 'SIN FAMILIA')", "Familia"),
                        'format': ("COALESCE(p.format, 'SIN FORMATO')", "Formato"),
                        'sku': ("COALESCE(oi.product_sku, 'SIN SKU')", "SKU Original")
                    }

                    group_field, group_label = group_field_map[group_by]
                    ws.title = f"Agrupado por {group_label}"

                    # Headers
                    headers = [group_label, "Pedidos", "Cantidad", "Revenue"]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border

                    # Query aggregated data
                    agg_query = f"""
                        SELECT
                            {group_field} as group_value,
                            COUNT(DISTINCT o.external_id) as pedidos,
                            SUM(oi.quantity) as cantidad,
                            SUM(oi.subtotal) as total_revenue
                        FROM orders o
                        LEFT JOIN order_items oi ON oi.order_id = o.id
                        LEFT JOIN products p ON p.sku = oi.product_sku
                        LEFT JOIN channels ch ON ch.id = o.channel_id
                        LEFT JOIN customers cust_direct ON cust_direct.id = o.customer_id AND cust_direct.source = o.source
                        LEFT JOIN LATERAL (
                            SELECT customer_external_id FROM customer_channel_rules ccr
                            WHERE ccr.channel_external_id::text = (
                                CASE WHEN o.customer_notes ~ '^\\s*\\{{'
                                THEN o.customer_notes::json->>'channel_id_relbase' ELSE NULL END
                            ) AND ccr.is_active = TRUE LIMIT 1
                        ) ccr_match ON true
                        LEFT JOIN customers cust_channel ON cust_channel.external_id = ccr_match.customer_external_id AND cust_channel.source = 'relbase'
                        WHERE {where_sql}
                        GROUP BY {group_field}
                        HAVING {group_field} IS NOT NULL
                        ORDER BY total_revenue DESC
                    """

                    cursor.execute(agg_query, params)
                    rows = cursor.fetchall()

                    for row_idx, row in enumerate(rows, 2):
                        ws.cell(row=row_idx, column=1, value=row['group_value']).border = thin_border
                        ws.cell(row=row_idx, column=2, value=row['pedidos']).border = thin_border
                        ws.cell(row=row_idx, column=3, value=row['cantidad']).border = thin_border
                        cell = ws.cell(row=row_idx, column=4, value=float(row['total_revenue'] or 0))
                        cell.border = thin_border
                        cell.number_format = number_format_currency

                    # Column widths
                    ws.column_dimensions['A'].width = 40
                    ws.column_dimensions['B'].width = 12
                    ws.column_dimensions['C'].width = 12
                    ws.column_dimensions['D'].width = 15

                # ===== DETAIL MODE EXPORT =====
                else:
                    ws.title = "Detalle Pedidos"

                    # Headers
                    headers = ["Pedido", "Fecha", "Cliente", "Canal", "SKU Original", "SKU Primario",
                               "Familia", "Producto", "Cantidad", "Unidades", "Precio Unit.", "Total"]
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border

                    # Handle not_in_catalog filter
                    if not_in_catalog:
                        cursor.execute(f"""
                            SELECT DISTINCT oi.product_sku as sku, oi.product_name, o.source
                            FROM orders o
                            LEFT JOIN order_items oi ON oi.order_id = o.id
                            LEFT JOIN channels ch ON ch.id = o.channel_id
                            LEFT JOIN customers cust_direct ON cust_direct.id = o.customer_id AND cust_direct.source = o.source
                            LEFT JOIN LATERAL (
                                SELECT customer_external_id FROM customer_channel_rules ccr
                                WHERE ccr.channel_external_id::text = (
                                    CASE WHEN o.customer_notes ~ '^\\s*\\{{'
                                    THEN o.customer_notes::json->>'channel_id_relbase' ELSE NULL END
                                ) AND ccr.is_active = TRUE LIMIT 1
                            ) ccr_match ON true
                            LEFT JOIN customers cust_channel ON cust_channel.external_id = ccr_match.customer_external_id AND cust_channel.source = 'relbase'
                            WHERE {where_sql} AND oi.product_sku IS NOT NULL AND oi.product_sku != ''
                        """, params)
                        all_skus = cursor.fetchall()

                        unmapped_skus_set = set()
                        for row in all_skus:
                            sku_val = row['sku']
                            product_name = row['product_name'] or ''
                            sku_source = row.get('source', '')
                            official_sku, _, _, _, _ = map_sku_with_quantity(
                                sku_val, product_name, product_catalog, catalog_skus, catalog_master_skus, sku_source
                            )
                            if not official_sku:
                                unmapped_skus_set.add(sku_val)

                        if unmapped_skus_set:
                            placeholders = ','.join(['%s'] * len(unmapped_skus_set))
                            where_clauses.append(f"oi.product_sku IN ({placeholders})")
                            params.extend(list(unmapped_skus_set))
                            where_sql = " AND ".join(where_clauses)

                    # Main query (no pagination limit for export)
                    query = f"""
                        SELECT
                            o.external_id as order_external_id,
                            o.order_date,
                            o.source as order_source,
                            COALESCE(cust_direct.name, cust_channel.name, 'SIN NOMBRE') as customer_name,
                            COALESCE(ch.name, 'SIN CANAL') as channel_name,
                            oi.product_sku as sku,
                            oi.product_name,
                            oi.quantity,
                            ROUND(oi.subtotal / NULLIF(oi.quantity, 0), 2) as unit_price,
                            oi.subtotal as item_subtotal,
                            p.category,
                            p.subfamily as family
                        FROM orders o
                        LEFT JOIN order_items oi ON oi.order_id = o.id
                        LEFT JOIN products p ON p.sku = oi.product_sku
                        LEFT JOIN channels ch ON ch.id = o.channel_id
                        LEFT JOIN customers cust_direct ON cust_direct.id = o.customer_id AND cust_direct.source = o.source
                        LEFT JOIN LATERAL (
                            SELECT customer_external_id FROM customer_channel_rules ccr
                            WHERE ccr.channel_external_id::text = (
                                CASE WHEN o.customer_notes ~ '^\\s*\\{{'
                                THEN o.customer_notes::json->>'channel_id_relbase' ELSE NULL END
                            ) AND ccr.is_active = TRUE LIMIT 1
                        ) ccr_match ON true
                        LEFT JOIN customers cust_channel ON cust_channel.external_id = ccr_match.customer_external_id AND cust_channel.source = 'relbase'
                        WHERE {where_sql}
                        ORDER BY o.order_date DESC, o.id, oi.id
                        LIMIT 100000
                    """

                    cursor.execute(query, params)
                    rows = cursor.fetchall()

                    service = get_product_catalog_service()

                    # ===== PACK EXPANSION for Export =====
                    # First, expand PACK products into component rows
                    processed_rows = []
                    for row in rows:
                        row_dict = dict(row)
                        sku = row_dict.get('sku', '') or ''

                        # Check if this is a PACK product
                        pack_mappings = get_pack_component_mappings(sku, cursor)

                        if pack_mappings:
                            # Expand PACK into component rows
                            component_rows = expand_pack_to_components(row_dict, pack_mappings, cursor)
                            processed_rows.extend(component_rows)
                        else:
                            processed_rows.append(row_dict)

                    # Now process and write all rows (including expanded components)
                    for row_idx, row in enumerate(processed_rows, 2):
                        row_sku = row.get('sku', '')
                        product_name = row.get('product_name', '')
                        order_source = row.get('order_source', '')
                        quantity = row.get('quantity', 0) or 0
                        is_pack_component = row.get('is_pack_component', False)

                        # For PACK components, use pre-computed values
                        if is_pack_component:
                            sku_prim = row.get('sku_primario', '')
                            unidades = quantity  # Already in units
                            familia = row.get('category', '')
                        else:
                            # Map SKU normally
                            official_sku, _, _, product_info, _ = map_sku_with_quantity(
                                row_sku, product_name, product_catalog, catalog_skus, catalog_master_skus, order_source
                            )

                            # Get enriched data
                            if official_sku:
                                sku_prim = get_sku_primario(official_sku, conversion_map)
                                unidades = calculate_units(row_sku, quantity, conversion_map, order_source)
                                familia = product_info.get('category', '') if product_info else row.get('category', '')
                            else:
                                sku_prim = None
                                unidades = quantity
                                familia = row.get('category', '')

                        # Format date
                        order_date = row.get('order_date')
                        date_str = order_date.strftime('%Y-%m-%d') if order_date else ''

                        # Write row
                        ws.cell(row=row_idx, column=1, value=row.get('order_external_id', '')).border = thin_border
                        ws.cell(row=row_idx, column=2, value=date_str).border = thin_border
                        ws.cell(row=row_idx, column=3, value=row.get('customer_name', '')).border = thin_border
                        ws.cell(row=row_idx, column=4, value=row.get('channel_name', '')).border = thin_border
                        ws.cell(row=row_idx, column=5, value=row_sku).border = thin_border
                        ws.cell(row=row_idx, column=6, value=sku_prim or '').border = thin_border
                        ws.cell(row=row_idx, column=7, value=familia or '').border = thin_border
                        ws.cell(row=row_idx, column=8, value=product_name).border = thin_border
                        ws.cell(row=row_idx, column=9, value=quantity).border = thin_border
                        ws.cell(row=row_idx, column=10, value=unidades).border = thin_border

                        cell_price = ws.cell(row=row_idx, column=11, value=float(row.get('unit_price') or 0))
                        cell_price.border = thin_border
                        cell_price.number_format = number_format_decimal

                        cell_total = ws.cell(row=row_idx, column=12, value=float(row.get('item_subtotal') or 0))
                        cell_total.border = thin_border
                        cell_total.number_format = number_format_currency

                    # Column widths
                    col_widths = [15, 12, 30, 20, 18, 18, 15, 35, 10, 10, 12, 12]
                    for col, width in enumerate(col_widths, 1):
                        ws.column_dimensions[get_column_letter(col)].width = width

                # Freeze header row
                ws.freeze_panes = 'A2'

                # Save to BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                # Generate filename
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                if group_by:
                    filename = f"auditoria_agrupado_{group_by}_{timestamp}.xlsx"
                else:
                    filename = f"auditoria_detalle_{timestamp}.xlsx"

                return StreamingResponse(
                    output,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename={filename}"}
                )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error exporting audit data: {str(e)}")
