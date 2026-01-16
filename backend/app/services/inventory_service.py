"""
Service for inventory management operations using direct psycopg2.
"""
import io
import csv
from typing import List, Dict, BinaryIO, Tuple
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.core.database import get_db_connection_dict_with_retry


class InventoryService:
    """Service for inventory business logic"""

    @staticmethod
    def get_product_by_sku(sku: str) -> Dict:
        """Get product by SKU from database"""
        conn = get_db_connection_dict_with_retry()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                SELECT id, sku, name, current_stock, min_stock, category
                FROM products
                WHERE sku = %s AND is_active = true
            """, (sku,))

            return cursor.fetchone()
        finally:
            cursor.close()
            conn.close()

    @staticmethod
    def update_product_stock(product_id: int, new_stock: int, old_stock: int, reason: str):
        """Update product stock and record movement"""
        conn = get_db_connection_dict_with_retry()
        cursor = conn.cursor()

        try:
            # Update product current_stock
            cursor.execute("""
                UPDATE products
                SET current_stock = %s, updated_at = %s
                WHERE id = %s
            """, (new_stock, datetime.utcnow(), product_id))

            # Record inventory movement
            quantity_change = new_stock - old_stock
            cursor.execute("""
            # NOTE: inventory_movements table removed in migration 029
            # INSERT INTO inventory_movements
            # (product_id, movement_type, quantity, stock_before, stock_after, reason, created_by, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                product_id,
                "adjustment",
                quantity_change,
                old_stock,
                new_stock,
                reason,
                "admin",
                datetime.utcnow()
            ))

            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            cursor.close()
            conn.close()

    @staticmethod
    def get_all_products_with_stock() -> List[Dict]:
        """Get all products with their current stock"""
        conn = get_db_connection_dict_with_retry()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                SELECT
                    id, sku, name, category, current_stock, min_stock,
                    package_type, units_per_master_box, master_box_sku, master_box_name, is_active
                FROM products
                WHERE is_active = true
                ORDER BY category, name
            """)

            return cursor.fetchall()
        finally:
            cursor.close()
            conn.close()

    @staticmethod
    def get_stock_summary() -> Dict:
        """Get summary of inventory status"""
        conn = get_db_connection_dict_with_retry()
        cursor = conn.cursor()

        try:
            cursor.execute("""
                SELECT
                    COUNT(*) as total_products,
                    COUNT(*) FILTER (WHERE current_stock > 0) as products_in_stock,
                    COUNT(*) FILTER (WHERE current_stock = 0) as out_of_stock,
                    COUNT(*) FILTER (WHERE current_stock <= min_stock AND current_stock > 0) as low_stock,
                    SUM(current_stock) as total_units,
                    SUM(current_stock * sale_price) as total_value
                FROM products
                WHERE is_active = true
            """)

            result = cursor.fetchone()
            return dict(result) if result else {}
        finally:
            cursor.close()
            conn.close()

    def generate_inventory_template(self) -> io.BytesIO:
        """Generate Excel template with all SKUs and current stock"""
        # Read the SKU master file
        csv_path = Path("/home/javier/Proyectos/Grana/Grana_Platform/public/Archivos_Compartidos/Codigos_Grana_Ingles.csv")
        df_skus = pd.read_csv(csv_path, encoding='utf-8')

        # Get all SKUs (both unit SKUs and master box SKUs)
        all_skus = []

        for _, row in df_skus.iterrows():
            # Add unit SKU
            all_skus.append({
                "SKU": row['SKU'],
                "Descripción": row['PRODUCTO'],
                "Categoría": row['CATEGORÍA'],
                "Tipo": "Unidad",
                "Tipo Envase": row['TIPO ENVASE UNID']
            })

            # Add master box SKU
            if pd.notna(row['SKU CAJA MÁSTER']) and row['SKU CAJA MÁSTER']:
                all_skus.append({
                    "SKU": row['SKU CAJA MÁSTER'],
                    "Descripción": row['NOMBRE CAJA MÁSTER'],
                    "Categoría": row['CATEGORÍA'],
                    "Tipo": "Caja Master",
                    "Tipo Envase": "CAJA MASTER"
                })

        # Get current stock from database
        products_in_db = self.get_all_products_with_stock()
        stock_map = {p['sku']: p['current_stock'] or 0 for p in products_in_db}

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario Grana"

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write headers
        headers = ["Artículo", "Descripción", "Categoría", "Tipo", "Cantidad Actual", "Nueva Cantidad", "Sub Empresa"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Write data
        for row_num, sku_data in enumerate(all_skus, 2):
            sku = sku_data['SKU']
            current_stock = stock_map.get(sku, 0)

            data = [
                sku,
                sku_data['Descripción'],
                sku_data['Categoría'],
                sku_data['Tipo'],
                current_stock,
                current_stock,  # Nueva cantidad (editable por usuario)
                "GRANA"
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')

                # Format numbers
                if col_num in [5, 6]:  # Cantidad columns
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.number_format = '#,##0'

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file

    def preview_inventory_file(self, file_content: bytes, filename: str) -> Dict:
        """
        Preview Excel file content WITHOUT updating database.
        Just reads and returns the data for display.
        """
        try:
            # Read WMS Excel file (header is in row 0)
            df_wms = pd.read_excel(io.BytesIO(file_content), sheet_name=0, header=0)

            # Handle WMS Excel format where first row contains headers
            if 'Articulo' in str(df_wms.iloc[0, 0]) or 'Encuentra - WMS' in str(df_wms.columns[0]):
                new_headers = df_wms.iloc[0].values
                df_wms = df_wms[1:]
                df_wms.columns = new_headers
                df_wms.reset_index(drop=True, inplace=True)

            # Find columns
            sku_col = None
            desc_col = None
            qty_col = None

            for col in df_wms.columns:
                col_lower = str(col).lower()
                if 'articulo' in col_lower or 'sku' in col_lower:
                    sku_col = col
                if 'descripcion' in col_lower or 'descripción' in col_lower:
                    desc_col = col
                if 'cantidad' in col_lower:
                    qty_col = col

            if sku_col is None:
                return {
                    "status": "error",
                    "message": f"No se pudo identificar columna de SKU. Columnas: {list(df_wms.columns)}"
                }

            # Extract rows
            rows = []
            for idx, row in df_wms.iterrows():
                sku = str(row[sku_col]).strip() if pd.notna(row[sku_col]) else ""
                desc = str(row[desc_col]).strip() if desc_col and pd.notna(row[desc_col]) else ""
                qty = row[qty_col] if qty_col and pd.notna(row[qty_col]) else 0

                # Skip header rows
                if not sku or sku.lower() in ['articulo', 'sku', 'artículo', 'nan']:
                    continue

                try:
                    qty = int(float(qty))
                except (ValueError, TypeError):
                    qty = 0

                rows.append({
                    "sku": sku,
                    "descripcion": desc,
                    "cantidad": qty
                })

            return {
                "status": "success",
                "filename": filename,
                "total_rows": len(rows),
                "columns": {
                    "sku": sku_col,
                    "descripcion": desc_col if desc_col else "N/A",
                    "cantidad": qty_col if qty_col else "N/A"
                },
                "rows": rows
            }

        except Exception as e:
            return {
                "status": "error",
                "message": f"Error leyendo Excel: {str(e)}"
            }

    def process_inventory_upload(self, file_content: bytes, filename: str) -> Dict:
        """
        Process uploaded WMS Excel file and update database.

        Key logic:
        1. Codigos_Grana_Ingles.csv = source of truth (160 unique SKUs)
        2. WMS Excel = current warehouse quantities
        3. If SKU not in WMS Excel → quantity = 0
        4. Update ALL SKUs in database based on this logic
        """
        try:
            # Read WMS Excel file (header is in row 0)
            df_wms = pd.read_excel(io.BytesIO(file_content), sheet_name=0, header=0)

            # Handle WMS Excel format where first row contains headers
            # Column names might be: 'Encuentra - WMS', 'Unnamed: 1', etc.
            # Actual headers are in the first data row
            if 'Articulo' in str(df_wms.iloc[0, 0]) or 'Encuentra - WMS' in str(df_wms.columns[0]):
                # First row is header, use it
                new_headers = df_wms.iloc[0].values
                df_wms = df_wms[1:]
                df_wms.columns = new_headers
                df_wms.reset_index(drop=True, inplace=True)

            # Find SKU and Quantity columns
            sku_col = None
            qty_col = None

            for col in df_wms.columns:
                col_lower = str(col).lower()
                if 'articulo' in col_lower or 'sku' in col_lower or col_lower.strip() == 'artículo':
                    sku_col = col
                if 'cantidad' in col_lower:
                    qty_col = col

            if sku_col is None or qty_col is None:
                return {
                    "status": "error",
                    "message": f"No se pudieron identificar las columnas de SKU y Cantidad en el Excel. Columnas encontradas: {list(df_wms.columns)}"
                }

            # Create dictionary of SKUs from WMS Excel with their quantities
            wms_quantities = {}
            for _, row in df_wms.iterrows():
                sku = str(row[sku_col]).strip() if pd.notna(row[sku_col]) else None
                qty = row[qty_col] if pd.notna(row[qty_col]) else 0

                if sku and sku.lower() not in ['articulo', 'sku', 'artículo', 'nan']:
                    try:
                        wms_quantities[sku] = int(float(qty))
                    except (ValueError, TypeError):
                        wms_quantities[sku] = 0

            # Load ALL SKUs from Codigos_Grana_Ingles.csv (source of truth)
            csv_path = Path("/home/javier/Proyectos/Grana/Grana_Platform/public/Archivos_Compartidos/Codigos_Grana_Ingles.csv")
            df_skus = pd.read_csv(csv_path, encoding='utf-8')

            # Collect all unique SKUs (both product SKUs and master box SKUs)
            all_source_skus = set()

            # Add product SKUs
            for sku in df_skus['SKU'].dropna():
                all_source_skus.add(str(sku).strip())

            # Add master box SKUs
            for sku in df_skus['SKU CAJA MÁSTER'].dropna():
                if str(sku).strip():  # Skip empty strings
                    all_source_skus.add(str(sku).strip())

            # Prepare updates for ALL SKUs from source of truth
            updates = []
            for sku in all_source_skus:
                # If SKU is in WMS Excel, use that quantity
                # If NOT in WMS Excel, quantity = 0 (not in warehouse)
                quantity = wms_quantities.get(sku, 0)
                updates.append((sku, quantity))

            if not updates:
                return {
                    "status": "error",
                    "message": "No se encontraron SKUs para actualizar"
                }

            # Perform bulk update
            results = self.bulk_update_stock(
                updates,
                f"Actualización masiva desde WMS: {filename} | SKUs en Excel: {len(wms_quantities)} | Total SKUs actualizados: {len(all_source_skus)}"
            )

            # Add additional context to results
            results["status"] = "success"
            results["message"] = f"Inventario completo actualizado desde WMS"
            results["context"] = {
                "skus_in_wms": len(wms_quantities),
                "skus_set_to_zero": len(all_source_skus) - len(wms_quantities),
                "total_skus_updated": len(all_source_skus)
            }

            return results

        except Exception as e:
            return {
                "status": "error",
                "message": f"Error procesando archivo WMS: {str(e)}"
            }

    def bulk_update_stock(self, updates: List[Tuple[str, int]], reason: str = "bulk_update") -> Dict:
        """Bulk update stock for multiple products"""
        results = {
            "success": [],
            "errors": [],
            "not_found": [],
            "summary": {
                "total": len(updates),
                "updated": 0,
                "failed": 0,
                "not_found": 0
            }
        }

        for sku, new_quantity in updates:
            try:
                product = self.get_product_by_sku(sku)

                if not product:
                    results["not_found"].append({
                        "sku": sku,
                        "reason": "Product not found in database"
                    })
                    results["summary"]["not_found"] += 1
                    continue

                old_stock = product['current_stock'] or 0

                # Update stock
                self.update_product_stock(
                    product_id=product['id'],
                    new_stock=new_quantity,
                    old_stock=old_stock,
                    reason=reason
                )

                results["success"].append({
                    "sku": sku,
                    "name": product['name'],
                    "old_stock": old_stock,
                    "new_stock": new_quantity,
                    "change": new_quantity - old_stock
                })
                results["summary"]["updated"] += 1

            except Exception as e:
                results["errors"].append({
                    "sku": sku,
                    "error": str(e)
                })
                results["summary"]["failed"] += 1

        return results
